"""
Converte o JSON gerado pelo convert_meu_dinheiro.py em um Excel
formatado compatível com a importação do Finance AI.
"""
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

JSON_IN = "C:/Users/ander/Downloads/finance-ai-import.json"
XLSX_OUT = "C:/Users/ander/Downloads/finance-ai-dados-anderson.xlsx"

data = json.load(open(JSON_IN, encoding="utf-8"))

wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="6366F1")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Mapa id conta/cartão → nome
acc_name = {a["id"]: a["name"] for a in data.get("accounts", [])}
card_name = {c["id"]: c["name"] for c in data.get("cards", [])}
cat_name = {c["id"]: c["name"] for c in data.get("categories", [])}

# Instruções
ws = wb.create_sheet("Instruções")
lines = [
    "📘 FINANCE AI — DADOS DE ANDERSON (EXPORTADO DO MEU DINHEIRO)",
    "",
    "Este arquivo já vem com seus dados do sistema anterior.",
    "Para importar no Finance AI:",
    "",
    "1. Abra: https://andersonassissouza.github.io/finance-ai/#/import",
    "2. Arraste este arquivo na área indicada",
    "3. Confira o preview",
    "4. Confirme a importação",
    "",
    "O que tem aqui:",
    f"• {len(data.get('accounts', []))} contas",
    f"• {len(data.get('cards', []))} cartões",
    f"• {len(data.get('transactions', []))} transações",
    "",
    "Para adicionar mais dados no futuro:",
    "• Use este arquivo e adicione linhas nas abas correspondentes",
    "• Ou baixe o modelo limpo em Importar → 'Baixar modelo Excel'",
]
for i, line in enumerate(lines, 1):
    ws.cell(i, 1, line)
ws.column_dimensions["A"].width = 80

def write_sheet(name, headers, rows, widths=None):
    ws = wb.create_sheet(name)
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            ws.cell(i, j, v)
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

# === CONTAS ===
write_sheet(
    "Contas",
    ["Nome", "Tipo", "Moeda", "Saldo inicial", "Cor", "Ícone"],
    [
        [a["name"],
         {"checking":"Corrente","savings":"Poupança","wallet":"Dinheiro","investment":"Investimento"}.get(a["type"], "Corrente"),
         a.get("currency", "BRL"),
         a.get("initial_balance", 0),
         a.get("color", "#6366f1"),
         a.get("icon", "🏦")]
        for a in data.get("accounts", [])
    ],
    [25, 14, 8, 14, 10, 6]
)

# === CARTÕES ===
write_sheet(
    "Cartões",
    ["Nome", "Limite", "Dia fechamento", "Dia vencimento", "Conta padrão", "Cor início", "Cor fim"],
    [
        [c["name"], c.get("limit", 0), c.get("closing_day", 1), c.get("due_day", 10),
         acc_name.get(c.get("default_account_id"), ""),
         c.get("color_start", "#6366f1"), c.get("color_end", "#ec4899")]
        for c in data.get("cards", [])
    ],
    [28, 10, 16, 16, 18, 12, 10]
)

# === TRANSAÇÕES ===
tx_rows = []
for t in sorted(data.get("transactions", []), key=lambda x: x.get("date", "")):
    tipo = {"income":"Receita","expense":"Despesa","transfer":"Transferência"}.get(t.get("type"), "")
    tx_rows.append([
        t.get("date", ""),
        t.get("description", ""),
        t.get("amount", 0),
        tipo,
        acc_name.get(t.get("account_id"), ""),
        card_name.get(t.get("card_id"), ""),
        t.get("category_id", ""),
        cat_name.get(t.get("category_id"), ""),
        ", ".join(t.get("tags", []) or []),
        t.get("notes", "")
    ])
write_sheet(
    "Transações",
    ["Data", "Descrição", "Valor", "Tipo", "Conta", "Cartão", "ID Categoria", "Categoria", "Tags", "Observações"],
    tx_rows,
    [12, 35, 12, 14, 18, 22, 22, 24, 18, 25]
)

# === METAS (vazio pra começar) ===
write_sheet(
    "Metas",
    ["Nome", "Valor alvo", "Valor atual", "Aporte mensal", "Data limite", "Ícone", "Cor"],
    [],
    [25, 12, 12, 14, 12, 6, 10]
)

# === RECORRÊNCIAS (vazio) ===
write_sheet(
    "Recorrências",
    ["Descrição", "Valor", "Categoria", "Conta", "Frequência", "Dia", "Data início", "Data fim"],
    [],
    [28, 10, 20, 18, 14, 6, 14, 14]
)

# === CATEGORIAS (referência) ===
write_sheet(
    "Categorias (ref.)",
    ["ID", "Nome", "Grupo"],
    [[c["id"], c["name"], c.get("group", "")] for c in data.get("categories", [])],
    [24, 28, 20]
)

wb.save(XLSX_OUT)
print(f"✅ Excel gerado: {XLSX_OUT}")
print(f"   {len(data.get('accounts', []))} contas")
print(f"   {len(data.get('cards', []))} cartões")
print(f"   {len(data.get('transactions', []))} transações")
import os
print(f"   {os.path.getsize(XLSX_OUT) / 1024:.1f} KB")
