"""
Conversor Meu Dinheiro → Finance AI
Lê os .xlsx exportados do Meu Dinheiro e gera um JSON importável
via Configurações → Exportar/Importar → Importar JSON
"""
import openpyxl, glob, json, uuid, re, sys, unicodedata
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

def norm(s):
    if not s: return ""
    s = str(s).lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn").strip()

def uid(prefix): return prefix + uuid.uuid4().hex[:12]
def iso(d):
    if not d: return ""
    if hasattr(d, "isoformat"): return d.isoformat().split("T")[0]
    s = str(d)
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", s)
    if m: return m.group(1)
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})", s)
    if m: return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return ""

# ============ CATEGORIES ============
# Categorias padrão do Finance AI + extras para o Meu Dinheiro
DEFAULT_CATS = [
    {"id":"cat_salary","name":"Salário","group":"Receitas","icon":"💰","color":"#10b981","type":"income"},
    {"id":"cat_freelance","name":"Freelance","group":"Receitas","icon":"💼","color":"#06b6d4","type":"income"},
    {"id":"cat_dividend","name":"Rendimentos","group":"Receitas","icon":"📈","color":"#0ea5e9","type":"income"},
    {"id":"cat_prolabore","name":"Pró-labore","group":"Receitas","icon":"💼","color":"#14b8a6","type":"income"},
    {"id":"cat_other_in","name":"Outras receitas","group":"Receitas","icon":"🪙","color":"#14b8a6","type":"income"},
    {"id":"cat_rent","name":"Aluguel/Prestação","group":"Moradia","icon":"🏠","color":"#f59e0b"},
    {"id":"cat_utilities","name":"Contas da casa","group":"Moradia","icon":"💡","color":"#eab308"},
    {"id":"cat_maintenance","name":"Manutenção doméstica","group":"Moradia","icon":"🔧","color":"#ca8a04"},
    {"id":"cat_internet","name":"Internet/TV/Celular","group":"Moradia","icon":"📡","color":"#f97316"},
    {"id":"cat_grocery","name":"Supermercado","group":"Alimentação","icon":"🛒","color":"#ef4444"},
    {"id":"cat_restaurant","name":"Restaurante","group":"Alimentação","icon":"🍽️","color":"#f43f5e"},
    {"id":"cat_delivery","name":"Delivery","group":"Alimentação","icon":"🛵","color":"#dc2626"},
    {"id":"cat_fuel","name":"Combustível","group":"Transporte","icon":"⛽","color":"#8b5cf6"},
    {"id":"cat_rideshare","name":"Uber/Táxi","group":"Transporte","icon":"🚕","color":"#a855f7"},
    {"id":"cat_transit","name":"Transporte/Estacionamento","group":"Transporte","icon":"🚌","color":"#7c3aed"},
    {"id":"cat_vehicle","name":"Manutenção veículo","group":"Transporte","icon":"🚗","color":"#6d28d9"},
    {"id":"cat_vehicle_tax","name":"IPVA/Seguro veículo","group":"Transporte","icon":"📄","color":"#5b21b6"},
    {"id":"cat_health","name":"Saúde","group":"Saúde","icon":"⚕️","color":"#22c55e"},
    {"id":"cat_pharmacy","name":"Farmácia","group":"Saúde","icon":"💊","color":"#16a34a"},
    {"id":"cat_exams","name":"Exames médicos","group":"Saúde","icon":"🩺","color":"#15803d"},
    {"id":"cat_gym","name":"Academia","group":"Saúde","icon":"🏋️","color":"#166534"},
    {"id":"cat_education","name":"Educação","group":"Educação","icon":"📚","color":"#3b82f6"},
    {"id":"cat_school","name":"Escola (mensalidade)","group":"Educação","icon":"🎒","color":"#2563eb"},
    {"id":"cat_streaming","name":"Streaming/Assinaturas","group":"Lazer","icon":"📺","color":"#ec4899"},
    {"id":"cat_travel","name":"Viagens","group":"Lazer","icon":"✈️","color":"#d946ef"},
    {"id":"cat_travel_food","name":"Alimentação em viagem","group":"Lazer","icon":"🍽️","color":"#c026d3"},
    {"id":"cat_entertainment","name":"Entretenimento","group":"Lazer","icon":"🎬","color":"#c026d3"},
    {"id":"cat_sports","name":"Esporte","group":"Lazer","icon":"⚽","color":"#a21caf"},
    {"id":"cat_gambling","name":"Apostas/Jogos","group":"Lazer","icon":"🎰","color":"#86198f"},
    {"id":"cat_shopping","name":"Compras","group":"Compras","icon":"🛍️","color":"#e11d48"},
    {"id":"cat_clothing","name":"Vestuário","group":"Compras","icon":"👕","color":"#be123c"},
    {"id":"cat_gifts","name":"Presentes","group":"Compras","icon":"🎁","color":"#9f1239"},
    {"id":"cat_electronics","name":"Eletrônicos","group":"Compras","icon":"📱","color":"#881337"},
    {"id":"cat_personal","name":"Cuidados pessoais","group":"Pessoal","icon":"💇","color":"#9333ea"},
    {"id":"cat_beauty","name":"Salão de beleza","group":"Pessoal","icon":"💅","color":"#7e22ce"},
    {"id":"cat_subscription","name":"Assinaturas","group":"Pessoal","icon":"🔁","color":"#6b21a8"},
    {"id":"cat_taxes","name":"Impostos/Taxas","group":"Financeiro","icon":"🧾","color":"#78716c"},
    {"id":"cat_pension","name":"Pensão alimentícia","group":"Financeiro","icon":"👨‍👧","color":"#64748b"},
    {"id":"cat_interest","name":"Juros/Multas","group":"Financeiro","icon":"⚠️","color":"#57534e"},
    {"id":"cat_invest_out","name":"Aporte investimento","group":"Financeiro","icon":"📊","color":"#0891b2","type":"invest"},
    {"id":"cat_pension_private","name":"Previdência privada","group":"Financeiro","icon":"🏦","color":"#0e7490","type":"invest"},
    {"id":"cat_debt","name":"Dívida/Financiamento","group":"Financeiro","icon":"📉","color":"#44403c"},
    {"id":"cat_transfer","name":"Transferência","group":"Transferência","icon":"🔄","color":"#64748b","type":"transfer"},
    {"id":"cat_card_payment","name":"Pagamento de cartão","group":"Transferência","icon":"💳","color":"#475569","type":"transfer"},
    {"id":"cat_other","name":"Outros","group":"Outros","icon":"📎","color":"#6b7280"},
]

CAT_BY_ID = {c["id"]: c for c in DEFAULT_CATS}

# Mapeamento (Categoria, Subcategoria) do Meu Dinheiro → id no Finance AI
def map_category(cat, sub, desc=""):
    cat_n = norm(cat); sub_n = norm(sub); desc_n = norm(desc)

    # RECEITAS
    if "pro-labore" in cat_n or "pro labore" in cat_n or "prolabore" in cat_n:
        return "cat_prolabore"
    if "salario" in sub_n or cat_n == "receitas" and "salari" in sub_n:
        return "cat_salary"
    if cat_n == "receitas" and "transferencia" in sub_n:
        return "cat_transfer"
    if cat_n == "receitas" or cat_n == "outras receitas":
        return "cat_other_in"

    # PAGAMENTO CARTÃO
    if "pagamento de cartao" in cat_n or cat_n == "pagamento de cartão":
        return "cat_card_payment"

    # MORADIA
    if cat_n == "moradia":
        if "energia" in sub_n: return "cat_utilities"
        if "agua" in sub_n or "gas" in sub_n: return "cat_utilities"
        if "manutencao" in sub_n: return "cat_maintenance"
        if "aluguel" in sub_n: return "cat_rent"
        if "condominio" in sub_n: return "cat_rent"
        if "iptu" in sub_n: return "cat_taxes"
        return "cat_utilities"

    # ALIMENTAÇÃO & CONSUMO
    if "consumo" in cat_n and "aliment" in sub_n:
        if "ifood" in desc_n or "rappi" in desc_n or "delivery" in desc_n:
            return "cat_delivery"
        return "cat_restaurant"
    if "aliment" in sub_n and "viagem" not in cat_n:
        if "ifood" in desc_n or "rappi" in desc_n: return "cat_delivery"
        return "cat_restaurant"
    if cat_n == "viagem" and "aliment" in sub_n:
        return "cat_travel_food"
    if "lazer" in cat_n and "aliment" in sub_n:
        return "cat_travel_food"
    if "supermerc" in sub_n or "mercado" in sub_n:
        return "cat_grocery"
    if cat_n == "consumo":
        if "telefone" in sub_n or "celular" in sub_n: return "cat_internet"
        if "assinatura" in sub_n: return "cat_streaming"
        if "transporte" in sub_n and "app" in sub_n: return "cat_rideshare"
        if "salao" in sub_n or "beleza" in sub_n: return "cat_beauty"
        if "presente" in sub_n: return "cat_gifts"
        if "eletronic" in sub_n: return "cat_electronics"
        if "ifood" in desc_n: return "cat_delivery"
        return "cat_shopping"

    # SAÚDE
    if cat_n == "saude":
        if "medicament" in sub_n: return "cat_pharmacy"
        if "exame" in sub_n: return "cat_exams"
        if "plano" in sub_n: return "cat_health"
        if "academia" in sub_n: return "cat_gym"
        return "cat_health"

    # EDUCAÇÃO
    if "educacao" in cat_n:
        if "mensalidade" in sub_n or "escola" in sub_n: return "cat_school"
        return "cat_education"

    # VEÍCULOS
    if cat_n == "veiculos":
        if "combust" in sub_n: return "cat_fuel"
        if "estaciona" in sub_n: return "cat_transit"
        if "manutencao" in sub_n: return "cat_vehicle"
        if "imposto" in sub_n or "ipva" in sub_n or "seguro" in sub_n: return "cat_vehicle_tax"
        return "cat_vehicle"

    # LAZER / VIAGEM
    if cat_n == "viagem":
        if "restaurante" in sub_n: return "cat_travel_food"
        return "cat_travel"
    if cat_n == "lazer":
        if "volei" in sub_n or "esporte" in sub_n: return "cat_sports"
        if "arena" in sub_n or "jogo" in sub_n: return "cat_entertainment"
        return "cat_entertainment"

    # VESTUÁRIO
    if "vestuario" in cat_n: return "cat_clothing"

    # INVESTIMENTOS
    if "investim" in cat_n:
        if "previd" in sub_n: return "cat_pension_private"
        return "cat_invest_out"

    # FINANCIAMENTOS / DÍVIDAS
    if "financiam" in cat_n or "emprest" in cat_n:
        return "cat_debt"

    # OUTRAS / APOSTAS
    if "outras despesas" in cat_n:
        if "apost" in sub_n: return "cat_gambling"
        return "cat_other"

    # PENSÃO
    if "pensao" in cat_n: return "cat_pension"

    # Default
    return "cat_other"

# ============ CARREGA DADOS ============
print("Carregando .xlsx...")
all_rows = []
seen = set()
for f in sorted(glob.glob("C:/Users/ander/Downloads/Meu_Dinheiro_2026*.xlsx")):
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb['Lançamentos']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for row in range(2, ws.max_row + 1):
        d = {h: ws.cell(row, i + 1).value for i, h in enumerate(headers)}
        k = "|".join(str(d.get(x, "")) for x in ["Tipo", "Data efetiva", "Data prevista", "Valor efetivo", "Valor previsto", "Descrição", "Conta"])
        if k in seen: continue
        seen.add(k)
        all_rows.append(d)
print(f"  {len(all_rows)} lançamentos únicos")

# ============ MONTA ACCOUNTS e CARDS ============
# Banco Inter (corrente)
acc_inter = {
    "id": uid("acc_"),
    "name": "Banco Inter",
    "type": "checking",
    "initial_balance": 0,
    "currency": "BRL",
    "color": "#ff7a00",
    "icon": "🏦",
    "include_in_net_worth": True,
    "created_at": datetime.now().isoformat()
}
# Dinheiro (carteira)
acc_cash = {
    "id": uid("acc_"),
    "name": "Dinheiro",
    "type": "wallet",
    "initial_balance": 0,
    "currency": "BRL",
    "color": "#10b981",
    "icon": "💵",
    "include_in_net_worth": True,
    "created_at": datetime.now().isoformat()
}
# Cartão Inter Pessoal
card_inter = {
    "id": uid("card_"),
    "name": "Cartão Inter Pessoal",
    "limit": 18810,
    "closing_day": 1,  # do .xlsx
    "due_day": 5,      # do .xlsx
    "color_start": "#ff7a00",
    "color_end": "#dc2626",
    "icon": "💳",
    "default_account_id": acc_inter["id"],
    "created_at": datetime.now().isoformat()
}

ACC_MAP = {
    "Banco Inter": acc_inter["id"],
    "Dinheiro": acc_cash["id"],
    "Cartão Inter Pessoal": None  # marca como card
}
CARD_MAP = {
    "Cartão Inter Pessoal": card_inter["id"]
}

# ============ CONVERTE LANÇAMENTOS ============
transactions = []
skipped = []

for r in all_rows:
    tipo = r.get('Tipo')
    data_ef = iso(r.get('Data efetiva'))
    data_prev = iso(r.get('Data prevista'))
    data = data_ef or data_prev
    if not data:
        skipped.append(("sem data", r))
        continue

    val_ef = r.get('Valor efetivo')
    val_prev = r.get('Valor previsto')
    valor = val_ef if val_ef not in (None, "") else val_prev
    try: valor = float(valor)
    except: skipped.append(("valor inválido", r)); continue

    descricao = (r.get('Descrição') or '').strip()
    conta = r.get('Conta') or ''
    conta_transf = r.get('Conta transferência') or ''
    contato = r.get('Contato') or ''
    centro = r.get('Centro') or ''

    # SALDO INICIAL — ignora (ou poderia setar initial_balance)
    if tipo == 'Saldo inicial':
        continue

    # PAGAMENTO DE CARTÃO → transferência Banco Inter → Cartão (reduz fatura)
    if tipo == 'Pagamento':
        # No Finance AI, pagamento de cartão é uma transação simples de entrada no cartão
        # e saída da conta. Como o cartão é representado por card_id nas transações,
        # cria-se apenas 1 transação negativa na conta de origem.
        # conta_transf é o "de onde vem" — Banco Inter
        from_account_id = ACC_MAP.get(conta_transf) or acc_inter["id"]
        transactions.append({
            "id": uid("tx_"),
            "date": data,
            "description": descricao or f"Pagamento fatura {conta}",
            "amount": -abs(valor),
            "account_id": from_account_id,
            "card_id": None,
            "category_id": "cat_card_payment",
            "type": "expense",
            "tags": [contato] if contato and contato != 'Sem contato' else [],
            "notes": centro,
            "created_at": datetime.now().isoformat()
        })
        continue

    # TRANSFERÊNCIA entre contas
    if tipo == 'Transferência':
        from_id = ACC_MAP.get(conta)
        to_id = ACC_MAP.get(conta_transf)
        if not from_id or not to_id:
            skipped.append(("transf inválida", r)); continue
        g = uid("grp_")
        transactions.append({
            "id": uid("tx_"), "group_id": g, "date": data,
            "description": descricao or "Transferência",
            "amount": -abs(valor), "account_id": from_id,
            "card_id": None, "category_id": "cat_transfer", "type": "transfer",
            "tags": [contato] if contato and contato != 'Sem contato' else [],
            "notes": centro, "created_at": datetime.now().isoformat()
        })
        transactions.append({
            "id": uid("tx_"), "group_id": g, "date": data,
            "description": descricao or "Transferência",
            "amount": abs(valor), "account_id": to_id,
            "card_id": None, "category_id": "cat_transfer", "type": "transfer",
            "tags": [contato] if contato and contato != 'Sem contato' else [],
            "notes": centro, "created_at": datetime.now().isoformat()
        })
        continue

    # DESPESA ou RECEITA
    tx_type = "income" if tipo == 'Receita' else "expense"
    if tx_type == "income" and valor < 0: valor = abs(valor)
    if tx_type == "expense" and valor > 0: valor = -abs(valor)

    # determina account_id / card_id
    # Para compras no cartão: account_id=null, card_id=xxx (não debita da conta)
    # Para transações na conta: account_id=xxx, card_id=null
    account_id = None
    card_id = None
    if conta == 'Cartão Inter Pessoal':
        card_id = CARD_MAP[conta]
        account_id = None  # gasto no cartão NÃO debita da conta direto
    elif conta in ACC_MAP:
        account_id = ACC_MAP[conta]
    else:
        account_id = acc_inter["id"]  # fallback

    category_id = map_category(r.get('Categoria'), r.get('Subcategoria'), descricao)

    # detecta parcela na descrição (ex: "DELL - 1/10 3/10")
    installment = None
    m = re.search(r'(\d+)/(\d+)', descricao)
    if m:
        installment = {"index": int(m.group(1)), "total": int(m.group(2))}

    tx = {
        "id": uid("tx_"),
        "date": data,
        "description": descricao or "(sem descrição)",
        "amount": round(valor, 2),
        "account_id": account_id,
        "card_id": card_id,
        "category_id": category_id,
        "type": tx_type,
        "tags": [contato] if contato and contato != 'Sem contato' else [],
        "notes": centro,
        "source": "meu_dinheiro",
        "created_at": datetime.now().isoformat()
    }
    if installment:
        tx["installment"] = installment
    transactions.append(tx)

print(f"  {len(transactions)} transações convertidas")
if skipped:
    print(f"  {len(skipped)} ignoradas")

# ============ SALDOS INICIAIS (se houver registros de "Saldo inicial") ============
# Pelo que vi, todos vinham com 0. Vou manter 0 mesmo.

# ============ MONTA JSON FINAL ============
output = {
    "accounts": [acc_inter, acc_cash],
    "cards": [card_inter],
    "transactions": transactions,
    "budgets": [],
    "goals": [],
    "debts": [],
    "investments": [],
    "rules": [],
    "recurrences": [],
    "cost_centers": [],
    "categories": DEFAULT_CATS,
    "alerts": [],
    "tags": [],
    "automations": [],
    "settings": {
        "currency": "BRL",
        "theme": "light",
        "first_name": "Anderson",
        "mode": "personal"
    }
}

out_path = "C:/Users/ander/Downloads/finance-ai-import.json"
with open(out_path, "w", encoding="utf-8") as fh:
    json.dump(output, fh, ensure_ascii=False, indent=2)

# ============ RELATÓRIO ============
from collections import Counter
cat_dist = Counter(t['category_id'] for t in transactions)
print(f"\n✅ JSON gerado: {out_path}")
print(f"   Contas: {len(output['accounts'])}")
print(f"   Cartões: {len(output['cards'])}")
print(f"   Transações: {len(transactions)}")

income_total = sum(t['amount'] for t in transactions if t['type'] == 'income')
expenses_in_account = abs(sum(t['amount'] for t in transactions if t['type'] == 'expense' and t['account_id'] and not t['card_id']))
expenses_on_card = abs(sum(t['amount'] for t in transactions if t['type'] == 'expense' and t['card_id']))
card_payments = abs(sum(t['amount'] for t in transactions if t['category_id'] == 'cat_card_payment'))
print(f"\n   💰 Receitas totais:         R$ {income_total:>12,.2f}")
print(f"   💸 Despesas na conta:       R$ {expenses_in_account:>12,.2f}  (inclui pagamento cartão)")
print(f"   💳 Gastos no cartão:        R$ {expenses_on_card:>12,.2f}  (já será compensado por pagamentos)")
print(f"     ↳ Pagamentos fatura:      R$ {card_payments:>12,.2f}  (debitado na conta)")
print(f"   🟰 Saldo efetivo em conta:  R$ {income_total - expenses_in_account:>12,.2f}  (aproximado)")
print(f"   💳 Fatura aberta estimada:  R$ {expenses_on_card - card_payments:>12,.2f}  (gastos não pagos ainda)")

print("\n   Distribuição por categoria (top 15):")
for cat, n in cat_dist.most_common(15):
    c = CAT_BY_ID.get(cat, {})
    print(f"     {n:4d} | {c.get('icon','?')} {c.get('name', cat)}")

dates = sorted(t['date'] for t in transactions if t['date'])
print(f"\n   📅 Período: {dates[0]} a {dates[-1]}")

# Calcula saldo virtual do Banco Inter (quanto precisaria ser o inicial para ficar em 0 hoje)
saldo_virtual_inter = sum(t['amount'] for t in transactions if t['account_id'] == acc_inter['id'])
print(f"\n   💡 Saldo virtual calculado do Banco Inter: R$ {saldo_virtual_inter:,.2f}")
print(f"      (fluxo líquido: receitas em conta - despesas em conta)")
print(f"      Se seu saldo REAL hoje for X, ajuste initial_balance = X - ({saldo_virtual_inter:.2f})")

size_kb = len(json.dumps(output)) / 1024
print(f"\n   📦 Tamanho: {size_kb:.1f} KB")
print(f"\n" + "="*60)
print(f"✅ PRONTO! Use este arquivo no Finance AI:")
print(f"   C:/Users/ander/Downloads/finance-ai-import.json")
print(f"\n📝 COMO IMPORTAR:")
print(f"   1. Abra: https://andersonassissouza.github.io/finance-ai/")
print(f"   2. Menu ⚙️ Configurações")
print(f"   3. 'Exportar / Importar' → clique em '⬆ Importar JSON'")
print(f"   4. Selecione o arquivo e confirme (substitui todos os dados)")
print(f"\n💡 AJUSTE POR FIM:")
print(f"   - Vá em Contas → Banco Inter → ⚙️ → ajuste Saldo inicial")
print(f"   - Recomendado: saldo real hoje + {abs(saldo_virtual_inter):.2f}")
